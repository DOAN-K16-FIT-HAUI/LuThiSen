from __future__ import annotations

import logging
import os
import time
from queue import Queue

import cv2
import numpy as np
import openpyxl
from common.settings import Settings
from config import camera_path
from config import img_logo_path
from PyQt5 import QtCore
from PyQt5 import QtGui
from PyQt5 import QtWidgets
from PyQt5.QtGui import QPixmap
from thread.Thread_callAPI import APICallerThread
from uis.main_window import Ui_MainWindow
# import folium

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
)


class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.setWindowTitle('Detect Hat, Helmet and Mask')

        self.create_queue()
        self.start_time = 0
        self.connect_signal()
        self.timer = QtCore.QTimer(self)
        self.place = 'Detect Hat, \nHelmet and Mask'
        self.level = ''
        # self.type = ""
        self.obj_center = ()
        self.current_obj_center = ()
        self.color = None

        self.camera_path = camera_path
        self.ui.le_cam_path.setText(self.camera_path)

        pixmap = QPixmap(img_logo_path)
        self.ui.label_avatar.setPixmap(pixmap)

        self.is_video = False
        self.setting = Settings()
        self.image_path = None
        self.test = 'test'

    def connect_signal(self):
        self.ui.btn_choose_path.clicked.connect(self.choose_path_img)
        self.ui.btn_start.clicked.connect(self.start_thread)
        self.ui.btn_link_cam.clicked.connect(self.link_cam)
        self.ui.btn_clear.clicked.connect(
            lambda: self.update_results(mode='clear'),
        )
        self.ui.btn_pause.clicked.connect(self.pause)

    def link_cam(self):
        self.ui.le_cam_path.setText(0)

    def choose_path_img(self):
        options = QtWidgets.QFileDialog.Options()
        base_dir = os.getcwd()
        default_dir = os.path.join(base_dir, r'resource/images')
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            'Choose File',
            default_dir,
            'All Supported Files (*.png *.jpg *.jpeg *.bmp *.gif *.mp4 *.avi *.mkv);;Image Files (*.png *.jpg *.jpeg *.bmp *.gif);;Video Files (*.mp4 *.avi *.mkv);;All Files (*)',
            options=options,
        )

        if file_path:
            self.image_path = file_path
            # Kiểm tra phần mở rộng
            _, ext = os.path.splitext(file_path)
            ext = ext.lower()
            self.is_video = ext in ['.mp4', '.avi', '.mkv']
            self.ui.le_cam_path.setText(file_path)
        else:
            self.image_path = None

    def create_queue(self):
        self.capture_queue = Queue()
        self.stream_queue = Queue()

    def start_thread(self):

        self.create_queue()
        self.create_threads()

    def run_image_processing(self):
        # Kiểm tra thread_callapi có tồn tại không
        if not hasattr(self, 'thread_callapi') or self.thread_callapi is None:
            self.thread_callapi = APICallerThread()
            self.thread_callapi.start()

        image = cv2.imread(self.image_path)
        if image is not None:
            self.response = self.thread_callapi.call_api(
                self.setting.host_ocr_service, image,
            )
            self.display_results(
                image=cv2.imread(
                    '/home/anodi108/Desktop/project/Do_An_Tot_Nghiep/DATN_LuThiSen/resource/data/cropped_outputs/aligned_card.png',
                ),
                classes=self.response.cls,
                course=self.response.course,
                born=self.response.date,
                HKTT=self.response.hktt,
                MSV=self.response.msv,
                name=self.response.name,
            )
            self.display_image_on_label(self.ui.label_input, image)

            self.save_to_excel(
                response=self.response,
                file_name='student_data.xlsx',
            )
        else:
            logging.warning('Không đọc được ảnh.')
            return

    def run_video_stream(self):
        self.video_capture = cv2.VideoCapture(self.image_path)
        self.timer.timeout.connect(self.read_video_frame)
        self.timer.start(30)  # 30 ms ~ 33 fps

    def pause(self):
        pass

    def create_threads(self):
        self.thread_callapi = APICallerThread()
        self.thread_callapi.start()

        if not hasattr(self, 'image_path') or not self.image_path:
            return

        if self.is_video:
            # Nếu là video → chạy xử lý luồng
            self.run_video_stream()
        else:
            # Nếu là ảnh → hiển thị ảnh
            self.run_image_processing()

    def display_results(self, image, name, born, HKTT, classes, course, MSV):
        self.update_results(mode='push')

        self.display_image_on_label(self.ui.label_img_result1, image)
        self.ui.label_name_result1.setText(name)
        self.ui.label_born_result1.setText(born)
        self.ui.label_HKTT_result1.setText(HKTT)
        self.ui.label_class_result1.setText(classes)
        self.ui.label_course_result1.setText(course)
        self.ui.label_MSV_result1.setText(MSV)

    def update_results(self, mode='push'):
        label_img_results = [
            self.ui.label_img_result1, self.ui.label_img_result2,
        ]
        label_name_results = [
            self.ui.label_name_result1, self.ui.label_name_result2,
        ]
        label_born_results = [
            self.ui.label_born_result1, self.ui.label_born_result2,
        ]
        label_HKTT_results = [
            self.ui.label_HKTT_result1, self.ui.label_HKTT_result2,
        ]
        label_class_results = [
            self.ui.label_class_result1, self.ui.label_class_result2,
        ]
        label_course_results = [
            self.ui.label_course_result1, self.ui.label_course_result2,
        ]
        label_MSV_results = [
            self.ui.label_MSV_result1, self.ui.label_MSV_result2,
        ]

        if mode == 'push':
            for i in range(1, 0, -1):
                label_img_results[i].clear()
                if label_img_results[i-1].pixmap() is not None:
                    label_img_results[i].setPixmap(
                        label_img_results[i-1].pixmap().copy(),
                    )
                label_img_results[i].setPixmap(label_img_results[i-1].pixmap())
                label_name_results[i].setText(label_name_results[i-1].text())
                label_born_results[i].setText(label_born_results[i-1].text())
                label_HKTT_results[i].setText(label_HKTT_results[i-1].text())
                label_class_results[i].setText(label_class_results[i-1].text())
                label_course_results[i].setText(
                    label_course_results[i-1].text(),
                )
                label_MSV_results[i].setText(label_MSV_results[i-1].text())
        elif mode == 'clear':
            for i in range(2):
                label_img_results[i].setPixmap(
                    QtGui.QPixmap('./resources/icons/folder_icon.png'),
                )
                label_name_results[i].setText('Họ tên')
                label_born_results[i].setText('Ngày sinh')
                label_HKTT_results[i].setText('HKTT')
                label_class_results[i].setText('Lớp')
                label_course_results[i].setText('Khóa học')
                label_MSV_results[i].setText('Mã SV')

    def display_image_on_label(self, ui_label, image):
        pixmap = QtGui.QImage(
            image.data, image.shape[1], image.shape[0], QtGui.QImage.Format.Format_RGB888,
        ).rgbSwapped()
        ui_label.clear()
        ui_label.setPixmap(QtGui.QPixmap.fromImage(pixmap))

    # def paintEvent(self, e):
    #     # self.obj_center = (0, 0)
    #     # if self.stream_queue.qsize() > 0:
    #     #     text, origin_image = self.stream_queue.get()
    #     #     image = origin_image.copy()
    #     #     warning_level = self.thread_work.event.warning_level
    #     #     if warning_level != "200":
    #     #         # print(warning_level)
    #     #         current_time = time.time()
    #     #         if (current_time - self.start_time) >= TIME_TO_PUSH_EVENT:
    #     #             self.display_results(image, self.place, text)
    #     #             self.save_img(image, img_result_path)
    #     #             # self.thread_client.send_message(warning_level)
    #     #             self.start_time = current_time
    #     #             self.thread_audio.play_audio(warning_level)
    #     #             # Hiệu ứng cảnh báo đỏ
    #     #             red_image = np.zeros(image.shape, image.dtype)
    #     #             red_image[:, :] = (0, 0, 255)
    #     #             red_factor = 0.2
    #     #             image = cv2.addWeighted(red_image, red_factor, image, 1 - red_factor, 0)
    #     #     self.display_image_on_label(self.ui.label_input,image)

    #     self.update()

    def read_video_frame(self):
        if self.video_capture.isOpened():
            ret, frame = self.video_capture.read()
            if ret:
                rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                self.display_image_on_label(self.ui.label_input, rgb_frame)
            else:
                self.timer.stop()
                self.video_capture.release()
                logging.info('Đã đọc hết video.')
        else:
            self.timer.stop()

    def save_img(self, image: np.ndarray, save_dir: str):
        timestamp = time.strftime('%Y%m%d_%H%M%S')
        filename = f'image_{timestamp}.jpg'
        save_path = os.path.join(save_dir, filename)
        cv2.imwrite(save_path, image)

    def send_data_to_api(self, api_url, data):
        self.thread_callapi.call_api(api_url, data)

    def save_to_excel(self, response, file_name='student_data.xlsx'):
        # Kiểm tra xem file đã tồn tại chưa
        try:
            # Nếu file đã tồn tại, mở nó
            wb = openpyxl.load_workbook(file_name)
            sheet = wb.active
        except FileNotFoundError:
            # Nếu file chưa tồn tại, tạo file mới
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = 'Dữ liệu'
            # Tiêu đề cột nếu là file mới
            sheet.append([
                'MSV', 'Tên', 'Khóa học',
                'Ngày sinh', 'HKTT', 'Lớp',
            ])

        # Kiểm tra xem MSV đã tồn tại trong file chưa
        msv_exists = False
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                if cell.value == response.msv:  # So sánh với MSV
                    msv_exists = True
                    break
            if msv_exists:
                break

        if msv_exists:
            print(
                f'MSV {response.msv} đã tồn tại trong file Excel. Không thêm dữ liệu mới.',
            )
        else:
            # Thêm dữ liệu vào hàng mới nếu MSV chưa có
            data = [
                response.msv,
                response.name,
                response.course,
                response.date,
                response.hktt,
                response.cls,
            ]
            sheet.append(data)

            # Đánh dấu theo MSV (Màu nền ô theo mã sinh viên)
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
                for cell in row:
                    if cell.value == response.msv:
                        cell.style = 'Good'  # Sử dụng phong cách "Good" của openpyxl

            # Lưu vào file
            wb.save(file_name)
            print(f'Dữ liệu đã được lưu vào {file_name}')
